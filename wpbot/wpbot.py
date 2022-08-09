import pyautogui as pg
import webbrowser as web
import time
import pandas as pd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

df1 = pd.read_excel('04.08sms.xlsx')
print(df1)


first = True
for i in range(len(df1['Phone'])):
    row = df1.loc[i]
    name = str(row[1]).lower()
    username = str(row[2]).lower()
 

    message = f" hi {name} {username}, your email address is {row[0]}"
    time.sleep(4)
    try:
        web.open("https://web.whatsapp.com/send?phone=+90"+str(row[3])+"&text="+message)
    except: 
        ws[f"A{i+1}"] = row[0]
        ws[f"B{i+1}"] = row[1]
        ws[f"C{i+1}"] = row[2]
        ws[f"D{i+1}"] = row[3]
    if first:
        time.sleep(6)
        first=False
    width,height = pg.size()
    pg.click(width/2,height/2)
    time.sleep(5)
    pg.press('enter')
    time.sleep(5)
    pg.hotkey('ctrl', 'w')


wb.save(f"mesaj gönderimi sırasında hata olanlar.xlsx")

